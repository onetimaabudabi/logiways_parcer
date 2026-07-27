"""Парсер MOHILL LINE — морские тарифы Азия ↔ Врангель + ЖД Врангель → РФ."""

from __future__ import annotations

import re
from datetime import datetime
from typing import Optional

import openpyxl

from .models import TariffSegment
from .utils import to_segments as _to_segments
from shared import port_dict, container_size_dict, get_country

_COMPANY = "Mohill Rus"
_VRANGEL = "Терминал Врангель"
_VRANGEL_CITY = "Находка"


def _parse_price(val) -> Optional[float]:
    if val is None:
        return None
    s = re.sub(r"[^\d]", "", str(val).replace("\xa0", ""))
    return float(s) if s else None


def _translate_port(raw: str) -> str:
    raw = re.sub(r"\*+", "", raw).strip()
    if raw in port_dict:
        return port_dict[raw]
    base = re.sub(r"\s*\(.*?\)", "", raw).strip()
    return port_dict.get(base, raw)


def _fmt_date(val) -> Optional[str]:
    if isinstance(val, datetime):
        return val.strftime("%Y-%m-%d")
    return None


def _parse_import_coc(rows: list, start_idx: int) -> list[dict]:
    """
    COC FILO section.
    rows[start_idx]   = main header: col7="COC, drop off Moscow" col9="...St.Peterburg" etc.
    rows[start_idx+1] = sub-header: col4=ETA, col5=ETD, col6=ETA, col7..14 = 20'/40' pairs
    rows[start_idx+2..] = data
    """
    results: list[dict] = []
    hdr = rows[start_idx] if start_idx < len(rows) else []

    # Detect drop-off destination columns from main header
    drop_cols: list[tuple[int, int, str]] = []  # (col20, col40, city_ru)
    for j, cell in enumerate(hdr):
        text = str(cell or "").strip().lower()
        if "moscow" in text:
            drop_cols.append((j, j + 1, "Москва"))
        elif "peterburg" in text or "petersburg" in text:
            drop_cols.append((j, j + 1, "Санкт-Петербург"))
        elif "ekaterinburg" in text:
            drop_cols.append((j, j + 1, "Екатеринбург"))
        elif "novosibirsk" in text:
            drop_cols.append((j, j + 1, "Новосибирск"))

    data_start = start_idx + 2
    for row in rows[data_start:]:
        if not row or not any(row):
            break
        pol_raw = str(row[0] or "").strip()
        # Stop at section separators
        if not pol_raw and not isinstance(row[5] if len(row) > 5 else None, datetime):
            break
        if not pol_raw:
            continue
        if pol_raw in ("POL", "MOHILL LINE") or "Ставки" in pol_raw or "ЖД" in pol_raw:
            break

        pol_ru = _translate_port(pol_raw)
        country = get_country(pol_ru) or "Китай"
        etd = _fmt_date(row[5]) if len(row) > 5 else None
        eta_pod = _fmt_date(row[6]) if len(row) > 6 else None
        vessel = str(row[2] or "").strip() or None
        voyage = str(row[3] or "").strip() or None
        departures_dict = {k: v for k, v in {"vessel": vessel, "voyage_no": voyage, "ETD": etd}.items() if v}

        for col20, col40, city_ru in drop_cols:
            for ct, col_idx in [("20DC", col20), ("40HC", col40)]:
                price = _parse_price(row[col_idx] if col_idx < len(row) else None)
                if price is None:
                    continue
                results.append(
                    TariffSegment(
                        transport_type="sea",
                        start_point=f"{pol_ru}, {country}",
                        end_point=f"{_VRANGEL}, Россия",
                        container_type=ct,
                        cost=price,
                        currency="USD",
                        company=_COMPANY,
                        container_ownership="COC",
                        port_service_term="FILO",
                        valid_from=etd,
                        valid_to=eta_pod,
                        start_location_type="port",
                        end_location_type="terminal",
                        parent_start_location=pol_ru,
                        parent_start_location_type="city",
                        parent_end_location=_VRANGEL_CITY,
                        parent_end_location_type="city",
                        departures=departures_dict,
                        weight_limit=container_size_dict.get(ct),
                        dropoff_location=city_ru,
                        dropoff_location_type="city",
                        dropoff_location_country="Россия",
                    ).to_dict()
                )
    return results


def _parse_import_soc(rows: list, start_idx: int) -> list[dict]:
    """
    SOC FILO section.
    rows[start_idx]   = main header: col7="SOC"
    rows[start_idx+1] = sub-header: col7=20', col8=40'
    rows[start_idx+2..] = data
    """
    results: list[dict] = []
    data_start = start_idx + 2
    for row in rows[data_start:]:
        if not row or not any(row):
            break
        pol_raw = str(row[0] or "").strip()
        if not pol_raw and not isinstance(row[5] if len(row) > 5 else None, datetime):
            break
        if not pol_raw:
            continue
        if pol_raw in ("POL", "MOHILL LINE") or "Ставки" in pol_raw or "ЖД" in pol_raw:
            break

        pol_ru = _translate_port(pol_raw)
        country = get_country(pol_ru) or "Китай"
        etd = _fmt_date(row[5]) if len(row) > 5 else None
        eta_pod = _fmt_date(row[6]) if len(row) > 6 else None
        vessel = str(row[2] or "").strip() or None
        voyage = str(row[3] or "").strip() or None
        departures_dict = {k: v for k, v in {"vessel": vessel, "voyage_no": voyage, "ETD": etd}.items() if v}

        for ct, col_idx in [("20DC", 7), ("40HC", 8)]:
            price = _parse_price(row[col_idx] if col_idx < len(row) else None)
            if price is None:
                continue
            results.append(
                TariffSegment(
                    transport_type="sea",
                    start_point=f"{pol_ru}, {country}",
                    end_point=f"{_VRANGEL}, Россия",
                    container_type=ct,
                    cost=price,
                    currency="USD",
                    company=_COMPANY,
                    container_ownership="SOC",
                    port_service_term="FILO",
                    valid_from=etd,
                    valid_to=eta_pod,
                    start_location_type="port",
                    end_location_type="terminal",
                    parent_start_location=pol_ru,
                    parent_start_location_type="city",
                    parent_end_location=_VRANGEL_CITY,
                    parent_end_location_type="city",
                    departures=departures_dict,
                    weight_limit=container_size_dict.get(ct),
                ).to_dict()
            )
    return results


def _parse_import_rail(rows: list, start_idx: int) -> list[dict]:
    """
    CY-FOR railway section.
    start_idx points to the "Направления" header row.
    Data rows: city | 20DC (до24т) | None | 20DC_28 (>24т) | None | 40HC | None | ВОХР20 | ВОХР40
    """
    results: list[dict] = []
    valid_from = "2026-06-01"
    valid_to = "2026-06-15"

    data_start = start_idx + 1
    for row in rows[data_start:]:
        if not row or not any(row):
            break
        city_raw = str(row[0] or "").strip()
        if not city_raw or "Ставки" in city_raw or "Направления" in city_raw:
            continue

        # Extract city name (before \n) and station(s) in parentheses
        city_line = city_raw.split("\n")[0].strip()
        station_match = re.search(r'\(([^)]+)\)', city_raw)
        stations = (
            [s.strip() for s in station_match.group(1).split(",")]
            if station_match else [city_line]
        )

        price_20 = _parse_price(row[1] if len(row) > 1 else None)
        price_20_28 = _parse_price(row[3] if len(row) > 3 else None)
        price_40 = _parse_price(row[5] if len(row) > 5 else None)

        for station in stations:
            # (container_type, price, min_weight_kg, max_weight_kg)
            rail_variants = [
                ("20DC", price_20,    None,  24),
                ("20DC", price_20_28, 24, None),
                ("40HC", price_40,    None,  28),
            ]
            for ct, price, min_kg, max_kg in rail_variants:
                if price is None:
                    continue
                results.append(
                    TariffSegment(
                        transport_type="rail",
                        start_point=f"{_VRANGEL}, Россия",
                        end_point=f"{station}, Россия",
                        container_type=ct,
                        cost=price,
                        currency="RUB",
                        company=_COMPANY,
                        valid_from=valid_from,
                        valid_to=valid_to,
                        start_location_type="terminal",
                        end_location_type="rail_station",
                        parent_start_location=_VRANGEL_CITY,
                        parent_start_location_type="city",
                        parent_end_location=city_line,
                        parent_end_location_type="city",
                        weight_limit=container_size_dict.get(ct),
                        min_weight_kg=min_kg,
                        max_weight_kg=max_kg,
                    ).to_dict()
                )
    return results


def _parse_export(wb) -> list[dict]:
    """Export sheet: Vrangel Bay → Asian ports, SOC LIFO."""
    ws = wb["Export"]
    results: list[dict] = []
    valid_from = "2026-06-01"
    valid_to = "2026-06-30"

    in_table = False
    for row in ws.iter_rows(min_row=1, values_only=True):
        if not row or not any(row):
            continue
        cell0 = str(row[0] or "").strip()

        if cell0 == "POL":
            in_table = True
            continue
        if not in_table:
            continue
        if not cell0 or cell0.startswith("*") or cell0.startswith("Subject"):
            continue

        pod_raw = str(row[1] or "").strip()
        if not pod_raw:
            continue

        pod_ru = _translate_port(pod_raw)
        country = get_country(pod_ru) or "Китай"
        price_20 = _parse_price(row[2] if len(row) > 2 else None)
        price_40 = _parse_price(row[3] if len(row) > 3 else None)

        for ct, price in [("20DC", price_20), ("40HC", price_40)]:
            if price is None:
                continue
            results.append(
                TariffSegment(
                    transport_type="sea",
                    start_point=f"{_VRANGEL}, Россия",
                    end_point=f"{pod_ru}, {country}",
                    container_type=ct,
                    cost=price,
                    currency="USD",
                    company=_COMPANY,
                    container_ownership="SOC",
                    port_service_term="LIFO",
                    valid_from=valid_from,
                    valid_to=valid_to,
                    start_location_type="terminal",
                    end_location_type="port",
                    parent_start_location=_VRANGEL_CITY,
                    parent_start_location_type="city",
                    parent_end_location=pod_ru,
                    parent_end_location_type="city",
                    weight_limit=container_size_dict.get(ct),
                ).to_dict()
            )
    return results


def _parse_all(file_path: str) -> list[dict]:
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # --- Import sheet ---
    ws_imp = wb["Import"]
    rows = list(ws_imp.iter_rows(min_row=1, values_only=True))

    mohill_markers: list[int] = []
    rail_header_idx: int = -1

    for i, row in enumerate(rows):
        cell0 = str(row[0] or "").strip()
        if cell0 == "MOHILL LINE":
            mohill_markers.append(i)
        if "CY-FOR" in cell0 or "ЖД отправка" in cell0:
            rail_header_idx = i + 1  # +1 skips the section label, data header is next row

    results: list[dict] = []

    if len(mohill_markers) >= 1:
        results += _parse_import_coc(rows, mohill_markers[0] + 1)
    if len(mohill_markers) >= 2:
        results += _parse_import_soc(rows, mohill_markers[1] + 1)
    if rail_header_idx >= 0:
        results += _parse_import_rail(rows, rail_header_idx)

    # --- Export sheet ---
    results += _parse_export(wb)

    return results


def parse(file_path: str) -> list[TariffSegment]:
    """Точка входа парсера MOHILL LINE."""
    return _to_segments(_parse_all(str(file_path)))


# Алиас для обратной совместимости с parsers/__init__.py
def parse_Mohill(file_path: str):
    return parse(file_path)
